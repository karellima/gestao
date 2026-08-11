import io
from dataclasses import dataclass
from zipfile import BadZipFile

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy.orm import Session

from app.models.product import Category, Product
from app.models.unit import Unit


class InvalidProductImportFile(ValueError):
    """Indica que o arquivo não é um workbook Excel suportado."""


@dataclass(frozen=True)
class ProductImportContext:
    category_index: int
    has_cost_price: bool
    has_stock: bool
    units_by_key: dict[str, int]
    categories_by_name: dict[str, list[Category]]


@dataclass(frozen=True)
class ProductImportResult:
    imported: int
    errors: list[str]


def import_products(contents: bytes, db: Session) -> ProductImportResult:
    workbook = _load_workbook(contents)
    try:
        sheet = workbook.active
        context = _build_context(sheet, db)
        imported, errors = _import_rows(sheet, context, db)
        db.commit()
        return ProductImportResult(imported=imported, errors=errors)
    finally:
        workbook.close()


def _load_workbook(contents: bytes):
    try:
        return openpyxl.load_workbook(io.BytesIO(contents))
    except (BadZipFile, InvalidFileException, OSError) as exc:
        raise InvalidProductImportFile from exc


def _build_context(sheet, db: Session) -> ProductImportContext:
    try:
        header_row = next(sheet.iter_rows())
    except StopIteration as exc:
        raise InvalidProductImportFile from exc

    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in header_row]
    has_cost_price = "preco custo" in headers or "preço custo" in headers
    has_stock = any("estoque" in header for header in headers)
    category_index = 5 if not has_cost_price else 6
    return ProductImportContext(
        category_index=category_index,
        has_cost_price=has_cost_price,
        has_stock=has_stock,
        units_by_key=_unit_map(db),
        categories_by_name=_category_map(db),
    )


def _unit_map(db: Session) -> dict[str, int]:
    units = db.query(Unit).all()
    by_key = {}
    for unit in units:
        by_key[unit.abbreviation.strip().lower()] = unit.id
        by_key[unit.name.strip().lower()] = unit.id
    return by_key


def _category_map(db: Session) -> dict[str, list[Category]]:
    categories_by_name: dict[str, list[Category]] = {}
    for category in db.query(Category).all():
        key = category.name.strip().lower()
        categories_by_name.setdefault(key, []).append(category)
    return categories_by_name


def _import_rows(sheet, context: ProductImportContext, db: Session) -> tuple[int, list[str]]:
    imported = 0
    errors = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            row_imported, error = _import_row(row, context, db)
        except (TypeError, ValueError) as exc:
            errors.append(f"Linha {row_number}: {exc!s}")
            continue
        if error:
            errors.append(f"Linha {row_number}: {error}")
        elif row_imported:
            imported += 1
    return imported, errors


def _import_row(row, context: ProductImportContext, db: Session) -> tuple[bool, str | None]:
    if not row or all(value is None for value in row):
        return False, None

    name = str(row[0]).strip() if row[0] else ""
    sku = str(row[1]).strip() if row[1] else ""
    if not name or not sku:
        return False, "Nome e SKU são obrigatórios"

    category_id, category_error = _resolve_category(row, context)
    if category_error:
        return False, category_error

    unit_id, unit_error = _resolve_unit(row, context)
    if unit_error:
        return False, unit_error

    data = _product_data(row, context, category_id, unit_id)
    _save_product(data, context, db, row)
    return True, None


def _save_product(data: dict, context: ProductImportContext, db: Session, row) -> None:
    existing = db.query(Product).filter(Product.sku == data["sku"]).first()
    if existing:
        for key, value in data.items():
            if value is not None:
                setattr(existing, key, value)
    else:
        if context.has_stock:
            stock_index = context.category_index + 3
            minimum_stock_index = context.category_index + 4
            data["current_stock"] = _number_at(row, stock_index)
            data["min_stock"] = _number_at(row, minimum_stock_index)
        db.add(Product(**data))


def _resolve_category(row, context: ProductImportContext) -> tuple[int | None, str | None]:
    category_index = context.category_index
    category_name = _text_at(row, category_index)
    subcategory_name = _text_at(row, category_index + 1)
    if not category_name and not subcategory_name:
        return None, None

    target_category = category_name.lower()
    target_subcategory = subcategory_name.lower()
    if target_subcategory:
        return _resolve_subcategory(context, target_category, target_subcategory, category_name, subcategory_name)
    return _resolve_parent_category(context, target_category, category_name)


def _resolve_subcategory(context, target_category, target_subcategory, category_name, subcategory_name):
    matches = _subcategory_matches(context, target_category, target_subcategory)
    if len(matches) == 1:
        return matches[0].id, None
    if not matches:
        parent = f" sob '{category_name}'" if category_name else ""
        return None, f"Subcategoria '{subcategory_name}'{parent} não encontrada"
    return None, f"Subcategoria '{subcategory_name}' é ambígua. Especifique também a categoria pai."


def _subcategory_matches(context, target_category, target_subcategory):
    candidates = context.categories_by_name.get(target_subcategory, [])
    if not target_category:
        return [category for category in candidates if category.parent_id is not None]
    parent_ids = {
        category.id
        for category in context.categories_by_name.get(target_category, [])
        if not category.parent_id
    }
    return [category for category in candidates if category.parent_id in parent_ids]


def _resolve_parent_category(context, target_category, category_name):
    parents = [
        category
        for category in context.categories_by_name.get(target_category, [])
        if not category.parent_id
    ]
    if len(parents) == 1:
        return parents[0].id, None
    return None, f"Categoria '{category_name}' não encontrada"


def _resolve_unit(row, context: ProductImportContext) -> tuple[int | None, str | None]:
    unit_index = context.category_index + 2
    unit_display = _text_at(row, unit_index)
    unit_value = unit_display.lower()
    if not unit_display:
        return None, None
    unit_id = context.units_by_key.get(unit_value)
    if unit_id:
        return unit_id, None
    return None, f"Unidade '{unit_display}' não encontrada"


def _product_data(row, context, category_id, unit_id) -> dict:
    return {
        "name": _text_at(row, 0),
        "sku": _text_at(row, 1),
        "description": _text_at(row, 2) or None,
        "barcode": _text_at(row, 3) or None,
        "price": _optional_number_at(row, 4),
        "cost_price": _optional_number_at(row, 5) if context.has_cost_price else None,
        "category_id": category_id,
        "unit_id": unit_id,
    }


def _text_at(row, index: int) -> str:
    return str(row[index]).strip() if len(row) > index and row[index] else ""


def _number_at(row, index: int) -> float:
    return float(row[index]) if len(row) > index and row[index] is not None else 0


def _optional_number_at(row, index: int) -> float | None:
    return float(row[index]) if len(row) > index and row[index] is not None else None
