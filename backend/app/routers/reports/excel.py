from io import BytesIO

import openpyxl
from fastapi import APIRouter, Body, Depends
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from app.utils.security import get_current_user

router = APIRouter()


class ExcelExportColumn(BaseModel):
    header: str
    width: int | None = 15


class ExcelExportRequest(BaseModel):
    title: str
    columns: list[ExcelExportColumn]
    rows: list[dict]
    filename: str | None = None


HEADER_FILL = PatternFill(start_color="14B8A6", end_color="14B8A6", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGNMENT = Alignment(vertical="center", wrap_text=True)


#: Caracteres que o Excel recusa num nome de aba. Sem a troca, um título com
#: uma barra (uma data "01/2026", por exemplo) derruba o endpoint em 500.
_SHEET_TITLE_INVALID = str.maketrans({c: "-" for c in "[]:*?/" + chr(92)})

#: O nome de arquivo entra num header HTTP entre aspas. Aspa, barra e quebra de
#: linha vindas do cliente sairiam do campo e reescreveriam a resposta.
_FILENAME_ALLOWED = " .,-_()[]"


def _safe_sheet_title(title: str) -> str:
    limpo = (title or "").translate(_SHEET_TITLE_INVALID).strip()
    return limpo[:31] or "Dados"


def _safe_filename(name: str) -> str:
    limpo = "".join(c for c in (name or "") if c.isalnum() or c in _FILENAME_ALLOWED).strip()
    return limpo[:100] or "relatorio"


def _apply_cell_style(cell, is_header=False):
    if is_header:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
    else:
        cell.font = BODY_FONT
        cell.alignment = BODY_ALIGNMENT
    cell.border = THIN_BORDER


@router.post("/export-excel")
def export_excel(
    payload: ExcelExportRequest = Body(...),
    _=Depends(get_current_user),
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _safe_sheet_title(payload.title)

    for col_idx, col in enumerate(payload.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col.header)
        _apply_cell_style(cell, is_header=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = col.width or 15

    for row_idx, row in enumerate(payload.rows, 2):
        for col_idx, col in enumerate(payload.columns, 1):
            value = row.get(col.header, "")
            # `rows` é JSON livre: uma lista ou um objeto aninhado faria o
            # openpyxl levantar ValueError no meio da planilha, virando 500.
            if isinstance(value, (dict, list, tuple, set)):
                value = str(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _apply_cell_style(cell, is_header=False)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    disposition_filename = f"{_safe_filename(payload.filename or payload.title)}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{disposition_filename}"'
        },
    )
