from calendar import monthrange
from datetime import datetime, timedelta
from io import BytesIO
from typing import Optional

import openpyxl
from fastapi import APIRouter, Body, Depends
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload

from app.database import get_db
from app.models.contact import Contact
from app.models.financial import Transaction
from app.models.financial_category import FinancialCategory
from app.models.product import Product
from app.models.stock import StockMovement
from app.models.user import User
from app.utils.security import get_current_user, is_admin_user, require_module, user_deposit_ids

router = APIRouter(prefix="/api/reports", tags=["Relatórios"])


def _is_admin(db: Session, user: User) -> bool:
    return is_admin_user(db, user)


@router.get("/dashboard")
def get_dashboard(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    _=Depends(require_module("dashboard")),
):
    product_query = db.query(Product).filter(Product.is_active == True)
    if not _is_admin(db, current_user):
        deposit_ids = user_deposit_ids(current_user)
        if deposit_ids:
            product_query = product_query.filter(Product.deposit_id.in_(deposit_ids))
        else:
            product_query = product_query.filter(False)

    total_products = product_query.count()
    low_stock = (
        product_query
        .filter(Product.current_stock <= Product.min_stock)
        .count()
    )
    total_contacts = db.query(Contact).filter(Contact.is_active == True).count()

    now = datetime.utcnow()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    month_start = today_start.replace(day=1)
    month_end = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
    next_7 = today_start + timedelta(days=7)

    receitas = (
        db.query(func.coalesce(func.sum(Transaction.amount), 0))
        .filter(Transaction.type == "receita", Transaction.date >= month_start, Transaction.date < month_end)
        .scalar()
    )
    despesas = (
        db.query(func.coalesce(func.sum(Transaction.amount), 0))
        .filter(Transaction.type == "despesa", Transaction.date >= month_start, Transaction.date < month_end)
        .scalar()
    )

    month_pending = lambda q: q.filter(
        Transaction.status.in_(["pendente", "pago_parcial"]),
        Transaction.due_date >= month_start,
        Transaction.due_date < month_end,
    )

    a_pagar = (
        month_pending(
            db.query(func.coalesce(func.sum(Transaction.amount), 0))
        )
        .filter(Transaction.type == "despesa")
        .scalar()
    )
    a_receber = (
        month_pending(
            db.query(func.coalesce(func.sum(Transaction.amount), 0))
        )
        .filter(Transaction.type == "receita")
        .scalar()
    )
    qtd_pendentes = month_pending(db.query(func.count(Transaction.id))).scalar()

    def serialize_txn(t):
        return {
            "id": t.id,
            "description": t.description,
            "amount": t.amount,
            "type": t.type,
            "due_date": t.due_date.isoformat() if t.due_date else None,
            "contact": t.contact.name if t.contact else None,
        }

    qry_txn = lambda q: q.options(
        joinedload(Transaction.financial_category),
        joinedload(Transaction.contact),
    )

    a_pagar_list = [
        serialize_txn(t)
        for t in month_pending(qry_txn(db.query(Transaction)))
        .filter(Transaction.type == "despesa")
        .order_by(Transaction.due_date.asc())
        .all()
    ]
    a_receber_list = [
        serialize_txn(t)
        for t in month_pending(qry_txn(db.query(Transaction)))
        .filter(Transaction.type == "receita")
        .order_by(Transaction.due_date.asc())
        .all()
    ]

    # Agrupamento por categoria
    despesas_cat = (
        db.query(FinancialCategory.name, func.sum(Transaction.amount))
        .join(Transaction, Transaction.financial_category_id == FinancialCategory.id)
        .filter(Transaction.type == "despesa", Transaction.date >= month_start)
        .group_by(FinancialCategory.name)
        .all()
    )
    receitas_cat = (
        db.query(FinancialCategory.name, func.sum(Transaction.amount))
        .join(Transaction, Transaction.financial_category_id == FinancialCategory.id)
        .filter(Transaction.type == "receita", Transaction.date >= month_start)
        .group_by(FinancialCategory.name)
        .all()
    )

    # Contas vencidas (não pagas/recebidas) - separadas por tipo
    base_overdue = lambda q: q.filter(
        Transaction.due_date < today_start,
        ~Transaction.status.in_(["pago", "recebido"]),
    )

    def overdue_stats(tipo):
        total = (
            base_overdue(db.query(func.coalesce(func.sum(Transaction.amount), 0)))
            .filter(Transaction.type == tipo)
            .scalar()
        )
        count = (
            base_overdue(db.query(func.count(Transaction.id)))
            .filter(Transaction.type == tipo)
            .scalar()
        )
        items = (
            base_overdue(
                db.query(Transaction).options(
                    joinedload(Transaction.financial_category),
                    joinedload(Transaction.contact),
                )
            )
            .filter(Transaction.type == tipo)
            .order_by(Transaction.due_date.asc())
            .all()
        )
        return {
            "count": count,
            "total": total,
            "list": [
                {
                    "id": t.id,
                    "description": t.description,
                    "amount": t.amount,
                    "type": t.type,
                    "due_date": t.due_date.isoformat() if t.due_date else None,
                    "contact": t.contact.name if t.contact else None,
                }
                for t in items
            ],
        }

    overdue_pagar = overdue_stats("despesa")
    overdue_receber = overdue_stats("receita")

    # Vencimentos nos próximos 7 dias (separados por tipo)
    def next_due_stats(tipo):
        q = db.query(Transaction).options(
            joinedload(Transaction.financial_category),
            joinedload(Transaction.contact),
        ).filter(
            Transaction.due_date >= today_start,
            Transaction.due_date <= next_7,
            ~Transaction.status.in_(["pago", "recebido"]),
            Transaction.type == tipo,
        )
        return {
            "count": q.count(),
            "total": q.with_entities(func.coalesce(func.sum(Transaction.amount), 0)).scalar(),
            "list": [serialize_txn(t) for t in q.order_by(Transaction.due_date.asc()).all()],
        }

    next_pagar = next_due_stats("despesa")
    next_receber = next_due_stats("receita")

    # Últimos lançamentos
    recent = (
        db.query(Transaction)
        .options(
            joinedload(Transaction.financial_category),
            joinedload(Transaction.contact),
        )
        .order_by(Transaction.created_at.desc())
        .limit(5)
        .all()
    )
    recent_transactions = [
        {
            "id": t.id,
            "description": t.description,
            "amount": t.amount,
            "type": t.type,
            "status": t.status,
            "date": t.date.isoformat() if t.date else None,
            "category": t.financial_category.name if t.financial_category else None,
            "contact": t.contact.name if t.contact else None,
        }
        for t in recent
    ]

    # Evolução mensal (últimos 6 meses)
    monthly_evolution = []
    for i in range(5, -1, -1):
        m = month_start.month - i
        y = month_start.year
        while m < 1:
            m += 12
            y -= 1
        while m > 12:
            m -= 12
            y += 1
        _, last_day = monthrange(y, m)
        m_start = datetime(y, m, 1)
        m_end = datetime(y, m, last_day, 23, 59, 59)
        m_rec = (
            db.query(func.coalesce(func.sum(Transaction.amount), 0))
            .filter(Transaction.type == "receita", Transaction.date.between(m_start, m_end))
            .scalar()
        )
        m_desp = (
            db.query(func.coalesce(func.sum(Transaction.amount), 0))
            .filter(Transaction.type == "despesa", Transaction.date.between(m_start, m_end))
            .scalar()
        )
        monthly_evolution.append({
            "mes": f"{y}-{m:02d}",
            "receitas": m_rec,
            "despesas": m_desp,
        })

    return {
        "total_products": total_products,
        "low_stock_products": low_stock,
        "total_contacts": total_contacts,
        "monthly_receitas": receitas,
        "monthly_despesas": despesas,
        "monthly_balance": receitas - despesas,
        "a_pagar": a_pagar,
        "a_receber": a_receber,
        "qtd_pendentes": qtd_pendentes,
        "a_pagar_list": a_pagar_list,
        "a_receber_list": a_receber_list,
        "despesas_por_categoria": dict(despesas_cat),
        "receitas_por_categoria": dict(receitas_cat),
        "overdue_pagar": overdue_pagar,
        "overdue_receber": overdue_receber,
        "next_pagar": next_pagar,
        "next_receber": next_receber,
        "next_due_total": next_pagar["total"] + next_receber["total"],
        "next_due_count": next_pagar["count"] + next_receber["count"],
        "next_due_list": next_pagar["list"] + next_receber["list"],
        "recent_transactions": recent_transactions,
        "monthly_evolution": monthly_evolution,
    }


@router.get("/financial-summary")
def get_financial_summary(
    start_date: datetime = None,
    end_date: datetime = None,
    db: Session = Depends(get_db),
    _=Depends(require_module("financial_reports")),
):
    if not start_date:
        start_date = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0)
    if not end_date:
        end_date = datetime.utcnow()

    receitas = (
        db.query(func.coalesce(func.sum(Transaction.amount), 0))
        .filter(Transaction.type == "receita", Transaction.date.between(start_date, end_date))
        .scalar()
    )
    despesas = (
        db.query(func.coalesce(func.sum(Transaction.amount), 0))
        .filter(Transaction.type == "despesa", Transaction.date.between(start_date, end_date))
        .scalar()
    )

    return {
        "periodo_inicio": start_date,
        "periodo_fim": end_date,
        "total_receitas": receitas,
        "total_despesas": despesas,
        "saldo": receitas - despesas,
    }


@router.get("/stock-movements-summary")
def get_stock_summary(
    days: int = 30,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    _=Depends(require_module("stock_reports")),
):
    since = datetime.utcnow() - timedelta(days=days)

    mov_query = db.query(StockMovement).filter(StockMovement.created_at >= since)
    if not _is_admin(db, current_user):
        deposit_ids = user_deposit_ids(current_user)
        if not deposit_ids:
            return {"periodo_dias": days, "total_entradas": 0, "total_saidas": 0}
        mov_query = mov_query.filter(StockMovement.deposit_id.in_(deposit_ids))

    entradas = (
        mov_query.filter(StockMovement.movement_type == "entrada")
        .with_entities(func.coalesce(func.sum(StockMovement.quantity), 0)).scalar()
    )
    saidas = (
        mov_query.filter(StockMovement.movement_type == "saida")
        .with_entities(func.coalesce(func.sum(StockMovement.quantity), 0)).scalar()
    )

    return {
        "periodo_dias": days,
        "total_entradas": entradas,
        "total_saidas": saidas,
    }


class ExcelExportColumn(BaseModel):
    header: str
    width: int | None = 15


class ExcelExportRequest(BaseModel):
    title: str
    columns: list[ExcelExportColumn]
    rows: list[dict]
    filename: str | None = None


HEADER_FILL = PatternFill(start_color="14B8A6", end_color="14B8A6", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGNMENT = Alignment(vertical="center", wrap_text=True)


#: Caracteres que o Excel recusa num nome de aba. Sem a troca, um título com
#: uma barra (uma data "01/2026", por exemplo) derruba o endpoint em 500.
_SHEET_TITLE_INVALID = str.maketrans({c: "-" for c in r"[]:*?/\\"})

#: O nome de arquivo entra num header HTTP entre aspas. Aspa, barra e quebra de
#: linha vindas do cliente sairiam do campo e reescreveriam a resposta.
_FILENAME_ALLOWED = " .,-_()[]"


def _safe_sheet_title(title: str) -> str:
    limpo = (title or "").translate(_SHEET_TITLE_INVALID).strip()
    return limpo[:31] or "Dados"


def _safe_filename(name: str) -> str:
    limpo = "".join(c for c in (name or "") if c.isalnum() or c in _FILENAME_ALLOWED).strip()
    return limpo[:100] or "relatorio"


def _apply_cell_style(cell, is_header=False):
    if is_header:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
    else:
        cell.font = BODY_FONT
        cell.alignment = BODY_ALIGNMENT
    cell.border = THIN_BORDER


@router.post("/export-excel")
def export_excel(
    payload: ExcelExportRequest = Body(...),
    _=Depends(get_current_user),
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _safe_sheet_title(payload.title)

    for col_idx, col in enumerate(payload.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col.header)
        _apply_cell_style(cell, is_header=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = col.width or 15

    for row_idx, row in enumerate(payload.rows, 2):
        for col_idx, col in enumerate(payload.columns, 1):
            value = row.get(col.header, "")
            # `rows` é JSON livre: uma lista ou um objeto aninhado faria o
            # openpyxl levantar ValueError no meio da planilha, virando 500.
            if isinstance(value, (dict, list, tuple, set)):
                value = str(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _apply_cell_style(cell, is_header=False)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    disposition_filename = f"{_safe_filename(payload.filename or payload.title)}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{disposition_filename}"'
        },
    )
